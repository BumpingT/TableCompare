"""
结果导出器模块
将对比结果导出为带样式的 Excel 文件
v3 — 纵向格式，每行一个字段变化，清晰易读
"""
import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from .comparator import DiffResult, STATUS_COL, KEY_COL, STATUS_ADDED, STATUS_DELETED, STATUS_MODIFIED, STATUS_SAME

logger = logging.getLogger('table_diff.exporter')

# 颜色定义
FILL_ADDED = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
FILL_DELETED = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
FILL_MODIFIED = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
FILL_CHANGED_CELL = PatternFill(start_color='F5CBA7', end_color='F5CBA7', fill_type='solid')
FILL_HEADER = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
FILL_SUBHEADER = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

FONT_HEADER = Font(color='FFFFFF', bold=True, size=11)
FONT_NORMAL = Font(color='1E293B', size=10)
FONT_BOLD = Font(color='1E293B', size=10, bold=True)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC'),
)


class ResultExporter:
    """结果导出器，v3 纵向格式"""

    STATUS_LABELS = {
        STATUS_ADDED: '新增',
        STATUS_DELETED: '删除',
        STATUS_MODIFIED: '修改',
        STATUS_SAME: '相同',
    }

    @staticmethod
    def to_excel(
        diff_result: DiffResult,
        output_path: str,
        only_diff: bool = False,
        export_columns: list[str] | None = None,
    ) -> str:
        """
        将对比结果导出为 Excel（纵向格式：每条字段变化占一行）

        Args:
            diff_result: 对比结果
            output_path: 输出文件路径
            only_diff: 是否只导出差异行
            export_columns: 要导出的列名列表，None 表示导出全部对比列

        Returns:
            导出文件的完整路径
        """
        df = diff_result.merged_df
        changed_cells = diff_result.changed_cells
        key_column = diff_result.key_column
        compare_columns = diff_result.compare_columns

        # 确定要导出的列
        if export_columns is not None:
            export_cols = [c for c in export_columns if c in compare_columns]
        else:
            export_cols = compare_columns[:]
        if not export_cols:
            export_cols = compare_columns[:]

        logger.info(f"开始导出: {len(df)}行 列={export_cols} -> {output_path}")

        # 过滤只显示差异行
        if only_diff and not df.empty:
            df = df[df[STATUS_COL] != STATUS_SAME].copy()

        if df.empty:
            logger.warning("没有数据可导出")
            raise ValueError("没有数据可导出")

        # --- 构建纵向数据 ---
        # 表头: 键值 | 状态 | 字段名 | 旧值 | 新值
        headers = [key_column or '键', '状态', '字段名', '旧值', '新值']

        rows_data = []
        for row_idx, (_, row) in enumerate(df.iterrows()):
            status = row[STATUS_COL]
            status_label = ResultExporter.STATUS_LABELS.get(status, status)
            key_val = row.get(KEY_COL, '')

            if status == STATUS_ADDED:
                for col in export_cols:
                    new_val = row.get(f'{col}_新', '')
                    new_str = str(new_val) if new_val is not None and new_val != '' else ''
                    rows_data.append([key_val, status_label, col, '', new_str])

            elif status == STATUS_DELETED:
                for col in export_cols:
                    old_val = row.get(f'{col}_旧', '')
                    old_str = str(old_val) if old_val is not None and old_val != '' else ''
                    rows_data.append([key_val, status_label, col, old_str, ''])

            elif status == STATUS_MODIFIED:
                source_idx = df.index[row_idx]
                for col in export_cols:
                    old_val = row.get(f'{col}_旧', '')
                    new_val = row.get(f'{col}_新', '')
                    old_str = str(old_val) if old_val is not None and old_val != '' else ''
                    new_str = str(new_val) if new_val is not None and new_val != '' else ''
                    rows_data.append([key_val, status_label, col, old_str, new_str])

            # status == same: 不导出（因为 only_diff 已经过滤了）

        # --- 写入 Excel ---
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="差异明细")

        ws.append(headers)
        for row_vals in rows_data:
            ws.append(row_vals)

        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        wb.save(output_path)

        # --- 应用样式 ---
        ResultExporter._apply_styles(output_path, headers, rows_data, changed_cells, diff_result.stats)

        logger.info(f"导出成功: {output_path} ({len(rows_data)}行)")
        return output_path

    @staticmethod
    def _apply_styles(
        output_path: str, headers: list, rows_data: list,
        changed_cells: dict, stats: dict,
    ):
        """应用样式"""
        from openpyxl import load_workbook

        wb = load_workbook(output_path)
        ws = wb.active
        total_rows = len(rows_data) + 1

        # 表头样式
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=ci)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER

        # 数据行样式
        for ri in range(2, total_rows + 1):
            if ri % 200 == 0:
                from PySide6.QtWidgets import QApplication
                QApplication.processEvents()

            status_cell = ws.cell(row=ri, column=2)  # 状态列
            status = str(status_cell.value or '')

            if status == '新增':
                row_fill = FILL_ADDED
            elif status == '删除':
                row_fill = FILL_DELETED
            elif status == '修改':
                row_fill = FILL_MODIFIED
            else:
                row_fill = None

            # 整行底色
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.font = FONT_NORMAL
                cell.alignment = ALIGN_LEFT
                if row_fill:
                    cell.fill = row_fill

            # 状态列加粗居中
            status_cell.font = FONT_BOLD
            status_cell.alignment = ALIGN_CENTER

            # 修改行：旧值/新值标橙色
            if status == '修改':
                old_cell = ws.cell(row=ri, column=4)
                new_cell = ws.cell(row=ri, column=5)
                # 如果新旧值不同，标橙色
                if str(old_cell.value or '') != str(new_cell.value or ''):
                    old_cell.fill = FILL_CHANGED_CELL
                    new_cell.fill = FILL_CHANGED_CELL

        # 列宽
        col_widths = [18, 8, 14, 30, 30]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 自动筛选
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{total_rows}"

        # 统计信息 sheet
        ws2 = wb.create_sheet(title="统计信息")
        stats_data = [
            ('指标', '数值'),
            ('初始版总行数', stats.get('total_old', 0)),
            ('修改版总行数', stats.get('total_new', 0)),
            ('新增行数', stats.get('added', 0)),
            ('删除行数', stats.get('deleted', 0)),
            ('修改行数', stats.get('modified', 0)),
            ('相同行数', stats.get('same', 0)),
            ('', ''),
            ('键列', stats.get('key_column', '')),
            ('导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        for ri, (label, value) in enumerate(stats_data, 1):
            cell_a = ws2.cell(row=ri, column=1, value=label)
            cell_b = ws2.cell(row=ri, column=2, value=value)
            if ri == 1:
                cell_a.font = FONT_HEADER
                cell_a.fill = FILL_HEADER
                cell_b.font = FONT_HEADER
                cell_b.fill = FILL_HEADER
            else:
                cell_a.font = FONT_BOLD
                cell_b.font = FONT_NORMAL

        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 15

        wb.save(output_path)
