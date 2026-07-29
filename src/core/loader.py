"""
文件加载器模块
负责加载 Excel (.xlsx/.xls) 和 CSV 文件为 pandas DataFrame
v3 — 增加多 Sheet 支持
"""
import os
import csv
import logging
import pandas as pd

logger = logging.getLogger('table_diff.loader')


class FileLoader:
    """文件加载器，支持 .xlsx, .xls, .csv 格式，支持多 Sheet"""

    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}

    @staticmethod
    def load(path: str, sheet_name: str | int | None = None) -> tuple[pd.DataFrame, dict]:
        """
        加载文件并返回 DataFrame 和元数据

        Args:
            path: 文件路径
            sheet_name: 要加载的 sheet 名称或索引（仅 Excel），None 表示第一个

        Returns:
            (DataFrame, metadata_dict)
            metadata_dict 包含: filename, rows, cols, columns, extension, sheet_name

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式不支持或内容为空
            Exception: 其他加载异常
        """
        if not os.path.exists(path):
            raise FileNotFoundError(f"文件未找到: {path}")

        ext = os.path.splitext(path)[1].lower()
        if ext not in FileLoader.SUPPORTED_EXTENSIONS:
            raise ValueError(f"不支持的文件格式 '{ext}'，仅支持: {', '.join(FileLoader.SUPPORTED_EXTENSIONS)}")

        filename = os.path.basename(path)
        logger.info(f"开始加载文件: {filename}, sheet={sheet_name}")

        try:
            if ext == '.csv':
                df = FileLoader._load_csv(path)
                loaded_sheet = None
            else:
                df, loaded_sheet = FileLoader._load_excel(path, sheet_name=sheet_name)
        except Exception as e:
            logger.error(f"文件加载失败: {filename}, 错误: {e}")
            raise

        if df.empty:
            raise ValueError(f"文件内容为空: {filename}")

        # 清理列名（去除首尾空格）
        df.columns = [str(col).strip() if col is not None else f"未命名列_{i}"
                      for i, col in enumerate(df.columns)]

        # 将列名中的空字符串替换为有意义的名称
        df.columns = [f"列_{i}" if col == '' else col for i, col in enumerate(df.columns)]

        metadata = {
            'filename': filename,
            'rows': len(df),
            'cols': len(df.columns),
            'columns': list(df.columns),
            'extension': ext,
            'sheet_name': loaded_sheet,
        }

        logger.info(f"文件加载成功: {filename}, sheet={loaded_sheet}, {metadata['rows']}行 x {metadata['cols']}列")
        return df, metadata

    @staticmethod
    def get_sheet_names(path: str) -> list[str]:
        """
        获取 Excel 文件的所有 sheet 名称

        Args:
            path: Excel 文件路径

        Returns:
            sheet 名称列表（非 Excel 文件返回空列表）
        """
        ext = os.path.splitext(path)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception as e:
            logger.warning(f"读取 sheet 名称失败: {e}")
            return []

    @staticmethod
    def _load_excel(path: str, sheet_name: str | int | None = None) -> tuple[pd.DataFrame, str | None]:
        """
        加载 Excel 文件

        Args:
            path: 文件路径
            sheet_name: sheet 名称或索引，None 表示第一个

        Returns:
            (DataFrame, actual_sheet_name)
        """
        if sheet_name is not None:
            try:
                df = pd.read_excel(path, sheet_name=sheet_name, dtype=str, engine='openpyxl')
            except Exception:
                logger.warning(f"Sheet '{sheet_name}' 不存在，回退到第一个 sheet")
                sheet_name = None

        if sheet_name is None:
            # 读取第一个 sheet
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            first_sheet = wb.sheetnames[0] if wb.sheetnames else None
            wb.close()

            if first_sheet is None:
                raise ValueError("Excel 文件没有 sheet")

            df = pd.read_excel(path, sheet_name=first_sheet, dtype=str, engine='openpyxl')

        # 将 NaN 替换为 None
        df = df.where(pd.notna(df), None)

        # 确定实际的 sheet 名称
        if isinstance(sheet_name, str):
            actual_sheet = sheet_name
        elif isinstance(sheet_name, int):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            names = wb.sheetnames
            wb.close()
            actual_sheet = names[sheet_name] if 0 <= sheet_name < len(names) else str(sheet_name)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            actual_sheet = wb.sheetnames[0] if wb.sheetnames else None
            wb.close()

        return df, actual_sheet

    @staticmethod
    def _load_csv(path: str) -> pd.DataFrame:
        """加载 CSV 文件，自动检测编码"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin-1']
        last_error = None

        for enc in encodings:
            try:
                # 先检测分隔符
                with open(path, 'r', encoding=enc) as f:
                    sample = f.read(8192)
                dialect = csv.Sniffer().sniff(sample, delimiters=[',', '\t', ';', '|'])
                delimiter = dialect.delimiter
                logger.info(f"CSV 编码: {enc}, 分隔符: '{delimiter}'")
                break
            except Exception as e:
                last_error = e
                continue
        else:
            # 所有编码都失败，尝试用 latin-1 兜底
            enc = 'latin-1'
            with open(path, 'r', encoding=enc) as f:
                sample = f.read(8192)
            dialect = csv.Sniffer().sniff(sample, delimiters=[',', '\t', ';', '|'])
            delimiter = dialect.delimiter
            logger.warning(f"CSV 编码检测失败，使用 latin-1, 分隔符: '{delimiter}'")

        try:
            df = pd.read_csv(path, encoding=enc, delimiter=delimiter,
                             dtype=str, keep_default_na=False)
            # 将 '' 替换为 None
            df = df.replace('', None)
            df = df.where(pd.notna(df), None)
            return df
        except Exception as e:
            logger.error(f"CSV 解析失败: {e}")
            raise

    @staticmethod
    def preview(df: pd.DataFrame, n: int = 10) -> pd.DataFrame:
        """返回前 n 行预览数据"""
        return df.head(n).copy()
